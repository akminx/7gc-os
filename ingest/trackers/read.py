"""Read the two internal workbooks into canonical rows.

These are *observations of what the tracker says*, not ledger truth. The
distinction matters: the tracker is a source document like any other, and one of
its stated totals disagrees with its own cells. A reader that silently trusted
the total would launder that disagreement into the ledger, so nothing here
computes a total — it records both what the sheet states and what its cells say,
and leaves the comparison to `reconcile.py`.

The workbooks are the fund's private case-study material and are not in the
repository. Every entry point takes a path, and the tests skip when it is
absent, exactly as the database tests skip without a DSN.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

#: Rows whose first cell starts with one of these are chrome, not data.
_NOT_A_HOLDING = ("7GC", "Internal", "TOTAL", "Marks", "Case study", "Fund I positions")


def _dec(value: object) -> Decimal | None:
    """A number the sheet stated, or None if the cell is not a number.

    `int`/`float` both arrive from openpyxl; the float is converted via `str` so
    a value that displayed as 1048515 does not become 1048514.99999999. INV-11 —
    a float amount is wrong before anyone reads it.
    """
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        # Excel stores pasted numbers as text more often than anyone expects.
        # Returning None here made the cell vanish from its column total while
        # the stated total still counted it — an understated column that either
        # reports a false disagreement or, if the footer is text too, nothing at
        # all. Only genuinely non-numeric text stays None.
        cleaned = value.strip().replace(",", "").replace("$", "")
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        try:
            return Decimal(cleaned)
        except (InvalidOperation, ValueError):
            return None
    return None


def _as_range(value: object) -> tuple[date, date] | None:
    """The earliest and latest day this cell could mean.

    The tracker writes dates as text, and three shapes appear:

      "10/10/2024"          one day          -> (that day, that day)
      "7/2020"              a month          -> (Jul 1, Jul 31)
      "2020 / 2021 / 2023"  three purchases  -> (Jan 1 2020, Dec 31 2023)

    Collapsing the last two to a single day would invent precision the source
    does not have, and the invented day decides whether a lot was held at a
    quarter boundary (INV-7). A range says exactly what is known, so a caller can
    answer "held at D" with yes, no, or *cannot determine from this cell*.
    """
    if isinstance(value, datetime):
        return (value.date(), value.date())
    if isinstance(value, date):
        return (value, value)
    if not isinstance(value, str):
        return None
    text = value.strip()
    try:
        exact = datetime.strptime(text, "%m/%d/%Y").date()
        return (exact, exact)
    except ValueError:
        pass
    try:  # month only
        m = datetime.strptime(text, "%m/%Y").date()
        last = date(m.year + m.month // 12, m.month % 12 + 1, 1) - timedelta(days=1)
        return (m, last)
    except ValueError:
        pass
    # Non-capturing: a capture group makes findall return "20" rather than "2020".
    years = [int(t) for t in re.findall(r"\b(?:19|20)\d{2}\b", text)]
    if years:
        return (date(min(years), 1, 1), date(max(years), 12, 31))
    return None


def _as_date(value: object) -> date | None:
    """The tracker writes dates as TEXT — "10/10/2024", not a date cell.

    Every `acquired` came back None until this was noticed, which silently broke
    "which tranche is most recent". Ambiguous forms are left as None rather than
    guessed: "2020 / 2021 / 2023" (three acquisitions) and "7/2020" (month only)
    are real entries, and inventing a day for them would fabricate precision the
    source does not have. The raw text is preserved on `acquired_text`.
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%m/%d/%Y").date()
        except ValueError:
            return None
    return None


@dataclass(frozen=True)
class TrackerMark:
    """One cell: what the tracker says a holding was worth at one period."""

    company: str
    period_label: str
    amount: Decimal | None
    #: The literal cell text when it is not a number — "Realized 5/20/24: ...",
    #: an em dash for a period after exit. Kept verbatim: an exit note is
    #: evidence about the position, and discarding it loses why a cell is blank.
    note: str | None = None


@dataclass(frozen=True)
class TrackerSheet:
    """One fund's grid, plus the totals the sheet asserts for itself."""

    fund_label: str
    period_labels: list[str]
    companies: list[str]
    cost_basis: dict[str, Decimal]
    marks: list[TrackerMark]
    #: What the TOTAL row states per period. Recorded, never trusted.
    stated_totals: dict[str, Decimal] = field(default_factory=dict)
    stated_cost_total: Decimal | None = None

    def amount(self, company: str, period: str) -> Decimal | None:
        for m in self.marks:
            if m.company == company and m.period_label == period:
                return m.amount
        return None

    def cells_total(self, period: str) -> Decimal:
        """What the column actually adds up to — computed, never read."""
        return sum(
            (m.amount for m in self.marks if m.period_label == period and m.amount is not None),
            Decimal(0),
        )


@dataclass(frozen=True)
class Tranche:
    """One row from the Master Investment Breakdown.

    Not every row is a purchase. Jackpocket carries an `Exit` row whose
    "Investment ($)" column holds 3,100,000 of *proceeds*, and treating that as
    cost produced a confident, entirely false reconciliation finding. INV-9 —
    cost basis is not fair value, and realisation proceeds are neither.
    """

    company: str
    kind: str
    investment: Decimal
    entry_valuation: str | None
    share_price: Decimal | None
    share_count: Decimal | None
    acquired: date | None
    #: What the Date column literally said, including forms no date can hold.
    acquired_text: str | None = None
    #: Earliest and latest day the cell could mean. Equal when it names one day.
    acquired_range: tuple[date, date] | None = None

    @property
    def acquired_is_exact(self) -> bool:
        r = self.acquired_range
        return r is not None and r[0] == r[1]

    def held_by(self, on: date) -> bool | None:
        """Was this acquired on or before `on`? None when the cell cannot say."""
        r = self.acquired_range
        if r is None:
            return None
        if r[1] <= on:
            return True
        if r[0] > on:
            return False
        return None

    @property
    def is_exit(self) -> bool:
        """Money coming back out. Its "Investment ($)" cell holds proceeds."""
        return self.kind.strip().lower().startswith(("exit", "realis", "realiz", "distrib"))

    @property
    def is_investment(self) -> bool:
        """Money going in — including instruments that are not plain equity.

        Matching `kind == "fund"` exactly was too narrow: The Mom Project's third
        row is `Fund (Conv. Note)`, a real 250,000 investment, and excluding it
        produced a confident false finding that the workbooks disagreed. Between
        an over-broad and an over-narrow rule the answer is neither — see
        `is_recognised`.
        """
        return self.kind.strip().lower().startswith("fund")

    @property
    def is_recognised(self) -> bool:
        """Did we understand this row at all?

        A row that is neither an investment nor an exit is not silently bucketed
        into one. Both of this reader's first two bugs were exactly that — a row
        quietly counted as something it was not — so an unfamiliar kind becomes a
        finding instead of a guess.
        """
        return self.is_investment or self.is_exit

    @property
    def implied_cost(self) -> Decimal | None:
        """price x count, for checking the sheet against itself."""
        if self.share_price is None or self.share_count is None:
            return None
        return self.share_price * self.share_count


def read_valuation_tracker(path: Path) -> list[TrackerSheet]:
    """Both grids from the valuation tracker workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets: list[TrackerSheet] = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        header = next((r for r in rows if r and r[0] == "Company"), None)
        if header is None:
            continue
        # Bind every period label to its OWN column index. Inferring the first
        # period column from `len(header) - len(periods)` silently shifted every
        # column when the header contained a blank or merged cell, so the grid
        # read cleanly against the wrong data.
        period_at = {str(c): i for i, c in enumerate(header) if c and i >= 3}
        periods = list(period_at)

        marks: list[TrackerMark] = []
        cost: dict[str, Decimal] = {}
        companies: list[str] = []
        for row in rows:
            name = row[0]
            if not isinstance(name, str) or row is header:
                continue
            if name.startswith(_NOT_A_HOLDING):
                continue
            companies.append(name)
            basis = _dec(row[2])
            if basis is not None:
                cost[name] = basis
            for label, index in period_at.items():
                cell = row[index] if index < len(row) else None
                amount = _dec(cell)
                note = str(cell) if amount is None and cell not in (None, "") else None
                marks.append(TrackerMark(name, label, amount, note))

        total_row = next(
            (r for r in rows if isinstance(r[0], str) and r[0].startswith("TOTAL")), None
        )
        stated: dict[str, Decimal] = {}
        stated_cost = None
        if total_row is not None:
            stated_cost = _dec(total_row[2])
            for label, index in period_at.items():
                value = _dec(total_row[index]) if index < len(total_row) else None
                if value is not None:
                    stated[label] = value

        sheets.append(
            TrackerSheet(
                fund_label=ws.title,
                period_labels=periods,
                companies=companies,
                cost_basis=cost,
                marks=marks,
                stated_totals=stated,
                stated_cost_total=stated_cost,
            )
        )
    return sheets


def read_master_breakdown(path: Path) -> list[Tranche]:
    """Every tranche across the per-company sheets.

    Sheets named `Fund I >>` / `Fund II >>` are section dividers with no rows;
    a company sheet is one whose header row starts with `Type`.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    tranches: list[Tranche] = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        header_at = next((i for i, r in enumerate(rows) if r and r[0] == "Type"), None)
        if header_at is None:
            continue
        # "Jackpocket (Realized)" and "Jio (Indirect)" carry their status in the
        # tab name; the company is the part before the parenthesis.
        company = ws.title.split(" (")[0].strip()
        for row in rows[header_at + 1 :]:
            kind = row[0]
            investment = _dec(row[1])
            if not isinstance(kind, str) or investment is None:
                continue  # documentation lines and footnotes
            tranches.append(
                Tranche(
                    company=company,
                    kind=kind,
                    investment=investment,
                    entry_valuation=str(row[2]) if row[2] is not None else None,
                    share_price=_dec(row[3]),
                    share_count=_dec(row[4]),
                    acquired=_as_date(row[5]),
                    acquired_text=str(row[5]) if row[5] is not None else None,
                    acquired_range=_as_range(row[5]),
                )
            )
    return tranches
