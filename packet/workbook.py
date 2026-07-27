"""The Excel workbook, and the CSVs beside it, from the same `Table` values.

Two things here are defensive rather than cosmetic.

**A quoted passage is data, not a formula.** openpyxl types any string beginning
with `=` as a formula, and this packet's longest column is verbatim document
text. A cited passage opening with `=` would be stored as an expression, so what
the auditor opens is `#NAME?` where the evidence used to be. Every string cell is
forced to the shared-string type, which is non-destructive — the alternative
mitigation, prefixing an apostrophe, edits the evidence to protect it.

**Money keeps its scale.** Amounts are written as `Decimal`, which openpyxl
serialises exactly, rather than converted to float on the way out. A packet that
re-quantises its own figures at the last step has undone the decimal policy the
schema enforces.
"""

from __future__ import annotations

import csv
import io
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from packet.tables import Cell, Table

#: Excel forbids these in a sheet name and caps it at 31 characters.
_FORBIDDEN = ":\\/?*[]"
_MONEY_FORMAT = "#,##0.00"
_MAX_WIDTH = 60
_MIN_WIDTH = 10


def sheet_name(title: str) -> str:
    cleaned = "".join(" " if c in _FORBIDDEN else c for c in title).strip()
    return (cleaned or "Sheet")[:31]


def cell_text(value: Cell) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def csv_bytes(table: Table) -> bytes:
    """A table as CSV. Data only — the note and footer are prose, not rows.

    Excel reads a UTF-8 CSV as the local code page unless it is given a BOM, and
    this file carries company names and quoted document text. The BOM is there so
    a double-click does not mangle the evidence.
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(table.headers)
    for row in table.rows:
        writer.writerow([cell_text(c) for c in row])
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def _write_cell(sheet: Worksheet, row_index: int, column_index: int, value: Cell) -> None:
    cell = sheet.cell(row=row_index, column=column_index, value=value)
    if isinstance(value, str):
        # Set after assignment: openpyxl decides the type inside the value
        # setter, and a leading '=' makes that decision 'formula'.
        cell.data_type = "s"
    elif isinstance(value, Decimal):
        cell.number_format = _MONEY_FORMAT
    cell.alignment = Alignment(vertical="top", wrap_text=isinstance(value, str) and len(value) > 60)


def _widths(sheet: Worksheet, table: Table) -> None:
    for column, header in enumerate(table.headers, start=1):
        longest = len(header)
        for row in table.rows:
            longest = max(longest, min(len(cell_text(row[column - 1])), _MAX_WIDTH))
        sheet.column_dimensions[get_column_letter(column)].width = max(
            _MIN_WIDTH, min(longest + 2, _MAX_WIDTH)
        )


def _render(sheet: Worksheet, table: Table) -> None:
    _write_cell(sheet, 1, 1, table.title)
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    _write_cell(sheet, 2, 1, table.note)
    header_row = 4
    for column, header in enumerate(table.headers, start=1):
        _write_cell(sheet, header_row, column, header)
        sheet.cell(row=header_row, column=column).font = Font(bold=True)
    for offset, row in enumerate(table.rows, start=1):
        for column, value in enumerate(row, start=1):
            _write_cell(sheet, header_row + offset, column, value)
    sheet.freeze_panes = f"A{header_row + 1}"
    if table.headers:
        sheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(table.headers))}{header_row + len(table.rows)}"
        )
    footer_at = header_row + len(table.rows) + 2
    for offset, line in enumerate(table.footer):
        _write_cell(sheet, footer_at + offset, 1, line)
        sheet.cell(row=footer_at + offset, column=1).font = Font(italic=True)
    _widths(sheet, table)


def workbook_bytes(tables: list[Table]) -> bytes:
    """Every table as one workbook, in the order given."""
    book = Workbook()
    default = book.active
    if default is not None:
        book.remove(default)
    used: set[str] = set()
    for table in tables:
        name = sheet_name(table.title)
        suffix = 2
        while name.casefold() in used:
            name = sheet_name(f"{table.title} {suffix}")
            suffix += 1
        used.add(name.casefold())
        _render(book.create_sheet(title=name), table)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def write(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(payload)
